#!/usr/bin/env python3
from pathlib import Path    # in order to check Path.cwd()
import openpyxl
import os
import re
from search_by_recent_credit import search_by_recent_credit, prev_payment_search
from Updateable_values import wb_list, ignore_list, DATE_COLUMN, BALANCE_COLUMN, CREDIT_COLUMN, write_dir, DESCRIPTION_COLUMN
from Month_class import Month
from colorama import init, Fore, Back, Style
from termcolor import colored
init()
import datetime
import sys

'''INSTRUCTIONS'''
'''See readme. If copying and pasting output to word, '<br>' can be replaced with the find and replace command, with "^m". This will cause page breaks to appear.

Make sure the files being loaded are the most recent files!

The program will search for months in 2020, according to the YEAR constant. Other constants 
designate which columns correspond to credit, date, and balance columns.

Some sheets are manually ignored, since the balance sheet excel files have extra non-balance sheets.'''


# Initial user input loop. Asks whether the user would like to check the most recent balance, or to check by month.
while True:
    print ("~" * 30)
    recent = input("Check most recent credit payment y/n? \n'Yes' does not provide the current balance. Entering 'no' will allow checking by month, which will provide the balance up to the entered month.\n")
    print ("~" * 30)
    if recent.lower() == 'y' or recent.lower() == 'yes':
        recent_balance_check = True
        break
    elif recent.lower() == 'n' or recent.lower() == 'no':
        recent_balance_check = False
        break
    else:
        print("Invalid entry")


if recent_balance_check:
    search_by_recent_credit()


def last_search(current_sheet, current_row, month):
    '''Given current sheet and the month name, recursively search for the last occurence of that month under the date column.
        Starting at current_row = 1, and then recursively adding to current_row.'''
    # Search each cell in column A with a loop, starting at current_row.
    for i in range(current_row, current_sheet.max_row):
        # First get the target month and subsequent month in datetime format.
        target_time_value = month.datetime_format
        next_month_value = month.one_month_later()
        match = False
        # Get the current cell in the date column.
        cell = current_sheet.cell(row=i, column = DATE_COLUMN)

        # Initialize cell_time_value to None if empty, or to the value in the cell.
        cell_time_value = None
        if cell.value != None and type(cell.value) == datetime.datetime:
            cell_time_value = cell.value
            # Find a match if subsequent cells are the same month or a previous month.
            # Done this way because sometimes months are entered out of order.
            # Will not match future months.
            if cell_time_value < target_time_value:
                match = True
            if cell_time_value >= target_time_value and cell_time_value < next_month_value:
                match = True

        # If the cell is not empty and the month matches, get the values and coordinates of the cell.
        if match == True: #found proper month:
            coord = cell.row
            val = cell.value
            # If not yet at the max row, call the search function again, starting at new row.
            if cell.row < current_sheet.max_row:
                coord2, val2 = last_search(current_sheet, cell.row + 1, month_checked)
            # if, after the search function is recursively called, new values are found, return them (base case, no new values, so these == None).
            if coord2 != None and val2 != None:
                return coord2, val2
            # otherwise return the originally found values.
            else:
                return coord, val
    # Return none if nothing found (will be base case, where coord2 = None and val2 = None).
    return None, None


# Get user input if checking by month.
while not recent_balance_check:
    month = input("What month is being checked for the balance? Enter the first three letters of any month.")
    try:
        month_checked = Month(month)
        break
    # Handle error if the Month constructor is given an invalid month as a value.
    except ValueError:
        print('Invalid month entry')
        continue

# If getting balance by input month, print out the latest balance and relevant info,
# corresponding to the last entry for that month or a previous month.
if not recent_balance_check:
    # Row in workbook in which data will be written, to be incremented for each string to be written.
    write_dir()
    writtenbook = openpyxl.Workbook()
    writesheet = writtenbook.active
    # row and column to be written to:
    write_row = 1
    write_column = 'A'

    # Also create a new sheet to store date in table format.
    table_sheet = writtenbook.create_sheet("Payment Table")
    column_width = 50
    table_sheet["A1"] = "TENANT SHEET"
    table_sheet.column_dimensions['A'].width = column_width
    table_sheet["B1"] = "FINAL BALANCE ENTRY DATED:"
    table_sheet.column_dimensions['B'].width = column_width
    table_sheet["C1"] = "PAYMENT DATE"
    table_sheet.column_dimensions['C'].width = column_width
    table_sheet["D1"] = "PREV PAYMENT, IF NO PAYMENT FOUND"
    table_sheet.column_dimensions['D'].width = column_width
    table_sheet["E1"] = "WRITTEN PAYMENT DESCRIPTION"
    table_sheet.column_dimensions['E'].width = column_width
    # row and column to be written to:
    table_row = 2
    table_column = 1
    table_letters = "0ABCDE"

    for wbIndex in range(len(wb_list)):
        #Load each workbook one by one, and change the working directory as well.
        wb = openpyxl.load_workbook(wb_list[wbIndex][0], data_only=True)
        os.chdir(wb_list[wbIndex][1])
        print ("\n")

        # for each sheet in the workbook, use last_search to find the last entry in the date column that 
        # corresponds to the input month, or to any previous months. Start searching at row 1.
        for sheet in wb.sheetnames:
            string3_exists = False
            string4_exists = False
            if sheet in ignore_list:
                continue
            current_sheet = wb[sheet]
            coord, val = last_search(current_sheet, 1, month_checked)
            print("active sheet = ", current_sheet)
            string1 = "active sheet = " + str(current_sheet)

            # If there is an entry print out the date.
            if coord != None:
                string3_exists = True
                try:
                    readable_date = datetime.datetime.strftime(val, '%B %d, %Y')
                except TypeError:
                    readable_date = val
                # print("cell = A" + str(coord))  #optional line to print out cell coordinate of date

                # Print out the balance, which is in the same row as the date.
                payment = current_sheet.cell(row=coord, column=CREDIT_COLUMN).value
                balance = current_sheet.cell(row=coord, column=BALANCE_COLUMN).value
                print(f"Final balance entry, dated {readable_date}: $" + str(balance))
                string2 = f"Final balance entry, dated {readable_date}: $" + str(balance)
                string2_abbr = f"{readable_date}: $" + str(balance)

                # The balance column should never be empty, so it should never == None.
                # On the other hand, the payment column is only filled if a payment has been made.
                if payment != None:
                    print("Payment of ${1} received on {0}".format(readable_date, payment))
                    string3 = "Payment of ${1} received on {0}".format(readable_date, payment)
                    try:
                        description = str(current_sheet.cell(row=coord, column=DESCRIPTION_COLUMN).value) + ", " + str(current_sheet.cell(row=coord - 1, column=DESCRIPTION_COLUMN).value)
                    except TypeError:
                        description = None
                else:
                    print(f"No payment listed for final balance entry on {readable_date}.")
                    string3 = f"No payment listed for final balance entry on {readable_date}."
                    prev_payment_coord, prev_payment_row = prev_payment_search(current_sheet, coord)
                    if prev_payment_coord != None:
                        prev_payment_date = current_sheet.cell(row=prev_payment_row, column=DATE_COLUMN).value
                        try:
                            description = str(current_sheet.cell(row=prev_payment_row, column=DESCRIPTION_COLUMN).value) + ", " + str(current_sheet.cell(row=prev_payment_row - 1, column=DESCRIPTION_COLUMN).value)
                        except TypeError:
                            description = None
                        if prev_payment_date != None:
                            try:
                                readable_prev_payment_date = datetime.datetime.strftime(prev_payment_date, '%B %d, %Y')
                                print(f"Previous payment entry listed as ${current_sheet[str(prev_payment_coord)].value} received on {readable_prev_payment_date}.")
                                string4_exists = True
                                string4 = f"Previous payment entry listed as ${current_sheet[str(prev_payment_coord)].value} received on {readable_prev_payment_date}."
                            except TypeError:
                                print((f"Previous payment entry listed as ${current_sheet[str(prev_payment_coord)].value} received on {prev_payment_date}."))
                                string4_exists = True
                                string4 = (f"Previous payment entry listed as ${current_sheet[str(prev_payment_coord)].value} received on {prev_payment_date}.")
                        else:
                            print("No previous payment found.")
                            string4_exists = True
                            string4 = "No previous payment found."
                    else:
                        print("No previous payment found.")
                        string4_exists = True
                        string4 = "No previous payment found."
            else:
                print(f"No entry for {month} or for previous months.")
                string2 = f"No entry for {month} or for previous months."
                string2_abbr = f"No entry for {month} or for previous months."

            print("")

            # Finally, write the retrieved data to a new workbook, stored in a directory specified
            # by write_dir(). The actual filepath can be found in Updateable_values.py.
            write_dir()
            writesheet[write_column + str(write_row)] = string1
            write_row += 1
            writesheet[write_column + str(write_row)] = string2
            write_row += 1
            if string3_exists:
                writesheet[write_column + str(write_row)] = string3
                write_row += 1
            if string4_exists:
                writesheet[write_column + str(write_row)] = string4
                write_row += 1
            write_row += 1

            # Write data to table in table_sheet.
            table_column = 1
            table_sheet[table_letters[table_column] + str(table_row)] = str(current_sheet)[11:]
            table_column = 2
            table_sheet[table_letters[table_column] + str(table_row)] = string2_abbr
            if string3_exists:
                table_column = 3
                table_sheet[table_letters[table_column] + str(table_row)] = string3
            if string4_exists:
                table_column = 4
                if "Previous payment entry listed as" in string4:
                    string4 = string4.split()
                    string4 = string4[5:]
                    string4 = " ".join(string4)
                table_sheet[table_letters[table_column] + str(table_row)] = string4
            table_column = 5
            if description != None:
                table_sheet[table_letters[table_column] + str(table_row)] = description
            table_row += 1

if not recent_balance_check:
    # Save the new excel file.
    present = datetime.datetime.now()
    present = datetime.datetime.strftime(present, "%B-%d-%Y")
    writtenbook.save("BALANCE DATA {0}.xlsx".format(present))
    print(f"\nExcel file '{present}.xlsx' written to {Path.cwd()}")
'''TODO'''
'''
Possibly implement writing to multiple file types (excel, text format, html).
Possibly print number of tenants checked.'''


while True:
    ex =input("type any key to exit")
    if ex != '':
        sys.exit()