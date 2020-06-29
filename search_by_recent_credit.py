import openpyxl
import colorama
from termcolor import colored
from Updateable_values import wb_list, CREDIT_COLUMN, ignore_list, DATE_COLUMN, BALANCE_COLUMN, DESCRIPTION_COLUMN, write_dir
import os
import datetime
from pathlib import Path

def most_recent_search(current_sheet):
    ''' A function that returns the coordinate of the most recent entry in the credit column, 'E', (column 5).'''

    # Search starting at the max row. Increment by -1 rows. There is no row 0, so stop at row 1.
    for i in range(current_sheet.max_row, 1, -1):
        cell = current_sheet.cell(row=i, column=CREDIT_COLUMN)
        #  Return the first non-empty cell found.
        if cell.value != None:
            return cell.coordinate
    return None

def prev_payment_search(current_sheet, row):
    '''Similar to most_recent_search() in that it returns the coordinate
    of the most recent entry in the credit column (column 5, or 'E'). 
    But this function starts from the input row instead of max row,
    increments by -1, and stops at row 1. '''
    for i in range(row, 1, -1):
        cell = current_sheet.cell(row=i, column=CREDIT_COLUMN)
        #  Return the first non-empty cell found.
        if cell.value != None:
            return cell.coordinate, cell.row
    return None, None


def search_by_recent_credit():
    ''' Load each workbook. For each workbook, print out information
    pertaining to the most recent entry in the credit column (column 'E', or 5).'''
    new_workbook = False
    write_dir()
    writtenbook = openpyxl.Workbook()
    recent_credit_sheet = writtenbook.active
    column_width = 50
    recent_credit_sheet["A1"] = "TENANT SHEET"
    recent_credit_sheet.column_dimensions['A'].width = column_width
    recent_credit_sheet["B1"] = "MOST RECENT PAYMENT DATE"
    recent_credit_sheet.column_dimensions['B'].width = column_width
    recent_credit_sheet["C1"] = "PAYMENT AMOUNT"
    recent_credit_sheet.column_dimensions['C'].width = column_width
    recent_credit_sheet["D1"] = "WRITTEN PAYMENT DESCRIPTION"
    recent_credit_sheet.column_dimensions['D'].width = column_width
    # row and column to be written to:
    table_row = 2
    table_column = 1
    table_letters = "0ABCD"

    for wbIndex in range(len(wb_list)):
        #Load each workbook one by one, and change the working directory as well.
        wb = openpyxl.load_workbook(wb_list[wbIndex][0], data_only=True)
        os.chdir(wb_list[wbIndex][1])

        # Print out company name at the top of each new workbook.
        if new_workbook == True:
            print("<br>")
            print("Company = " + wb_list[wbIndex][3] + ", ", end="")
            print("Property = " + wb_list[wbIndex][2])
            print("~" * 80)

        if new_workbook == False:
            print("Company = " + wb_list[wbIndex][3] + ", ", end="")
            print("Property = " + wb_list[wbIndex][2])
            print("~" * 80)
            new_workbook = True

        for sheet in wb.sheetnames:
            # ignore the security deposit sheets
            if sheet in ignore_list:
                continue
            if most_recent_search(wb[sheet]) == None:
                print("No credit entries listed on", sheet, ".")
                continue
            credit_cell = wb[sheet][most_recent_search(wb[sheet])]
            # wb[sheet] is the active sheet. most_recent_search(wb[sheet]) returns a cell coordinate.
            # Print the name of the tenant, which correspondes to the current sheetname.
            print("")
            print("Tenant name = ", sheet)
            try:
                credit_date = datetime.datetime.strftime(wb[sheet].cell(row = credit_cell.row, column = DATE_COLUMN).value, '%B %d, %Y')
                print("Most recent payment = $", credit_cell.value, "listed on", credit_date + ".")
            except TypeError:
                print("Most recent payment = $", credit_cell.value, "No date Listed.")
            try:
                description = str(wb[sheet].cell(row = credit_cell.row, column = DESCRIPTION_COLUMN).value) + ", " + str(wb[sheet].cell(row = credit_cell.row - 1, column = DESCRIPTION_COLUMN).value)
            except TypeError:
                description = None

            write_dir()
            table_column = 1
            recent_credit_sheet[table_letters[table_column] + str(table_row)] = str(wb[sheet])
            table_column = 2
            try:
                recent_credit_sheet[table_letters[table_column] + str(table_row)] = credit_date
            except ValueError:
                pass
            table_column = 3
            recent_credit_sheet[table_letters[table_column] + str(table_row)] = credit_cell.value
            table_column = 4
            if description != None:
                recent_credit_sheet[table_letters[table_column] + str(table_row)] = description
            table_row += 1

    write_dir()
    present = datetime.datetime.now()
    present = datetime.datetime.strftime(present, "%B-%d-%Y")
    writtenbook.save("MOST RECENT PAYMENTS {0}.xlsx".format(present))
    print(f"\nExcel file '{present}.xlsx' written to {Path.cwd()}")