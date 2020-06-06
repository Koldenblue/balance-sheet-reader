import openpyxl, os
import re
from Workbook import Workbook

wb = openpyxl.load_workbook(r'\\Optiplex7440\c\Rents\Rent 2020\Tenant Rent\Marlin Westwood Tenant Balance Sheets/1736 Westwood Tenant Balance Sheets.xlsx', data_only=True)
os.chdir(r'\\Optiplex7440\c\Rents\Rent 2020\Tenant Rent\Marlin Westwood Tenant Balance Sheets')
print(wb.sheetnames)

while True:
    month = input("What month is being checked for the balance? Enter the first three letters of any month.")
    if month[0:3].lower() == 'jan':
        month_index = '01'
    elif month[0:3].lower() == 'feb':
        month_index = '02'
    elif month[0:3].lower() == 'mar':
        month_index = '03'
    elif month[0:3].lower() == 'apr':
        month_index = '04'
    elif month[0:3].lower() == 'may':
        month_index = '05'
    elif month[0:3].lower() == 'jun':
        month_index = '06'
    elif month[0:3].lower() == 'jul':
        month_index = '07'
    elif month[0:3].lower() == 'aug':
        month_index = '08'
    elif month[0:3].lower() == 'sep':
        month_index = '09'
    elif month[0:3].lower() == 'oct':
        month_index = '10'
    elif month[0:3].lower() == 'nov':
        month_index = '11'
    elif month[0:3].lower() == 'dec':
        month_index = '12'
    else:
        print("Invalid entry.")
        continue
    break

# Make a regex consisting of the month number, so that the excel sheet 'datetime.datetime' entries can be searched for that month using the regex.
month_regex = re.compile(f'-{month_index}-')

def last_search(current_sheet, current_row):
    '''Given current sheet and the month name, recursively search for the last occurence of that month under the date column.
        Starting at current_row = 1.'''
    # Search each cell in column A for the month match, starting at current_row.
    for i in range(current_row, current_sheet.max_row):
        cell = current_sheet.cell(row=i, column = 1)
        mo = month_regex.search(str(cell.value))
        # If the cell is not empty and the month matches, get the values and coordinates of the cell.
        if cell.value != None and mo != None:
            coord = cell.row
            val = cell.value
            # If not yet at the max row, call the search function again, starting at new row.
            if cell.row < current_sheet.max_row:
                coord2, val2 = last_search(current_sheet, cell.row + 1)
            # if, after the search function is recursively called, new values are found, return them (base case, no new values, so these == None).
            if coord2 != None and val2 != None:
                return coord2, val2
            # otherwise return the originally found values.
            else:
                return coord, val
    # Return none if nothing found (will be base case, where coord2 = None and val2 = None).
    return None, None

for sheet in wb.sheetnames:
    current_sheet = wb[sheet]
    coord, val = last_search(current_sheet, 1)
    print("active sheet = ", current_sheet)
    print("row = " + str(coord))
    print(val)
    
    balance = current_sheet.cell(row=coord, column=6).value
    print(balance)


    print ("\n")

        #if cell.value != None and type(cell.value) != 'datetime.datetime':
        #    print("Convert to date format!")
    # datetime format explained further in Ch 17 of how to automate

    #print(wb.active['A4'].value)
